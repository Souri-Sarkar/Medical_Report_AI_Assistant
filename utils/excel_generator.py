from openpyxl import Workbook
from openpyxl.styles import Font


def generate_excel(
    patient_name,
    age,
    gender,
    report_type,
    analysis,
    filename="data/output_reports/Medical_Report.xlsx"
):

    wb = Workbook()

    ws = wb.active
    ws.title = "Medical Report"

    # ---------------------------------------
    # Title
    # ---------------------------------------

    ws["A1"] = "Medical Report AI Assistant"
    ws["A1"].font = Font(
        bold=True,
        size=16
    )

    # ---------------------------------------
    # Patient Information
    # ---------------------------------------

    ws["A3"] = "Patient Name"
    ws["B3"] = patient_name

    ws["A4"] = "Age"
    ws["B4"] = age

    ws["A5"] = "Gender"
    ws["B5"] = gender

    ws["A6"] = "Report Type"
    ws["B6"] = report_type

    # ---------------------------------------
    # Analysis Header
    # ---------------------------------------

    row = 8

    ws.cell(row=row, column=1).value = "Parameter"
    ws.cell(row=row, column=2).value = "Value"
    ws.cell(row=row, column=3).value = "Unit"
    ws.cell(row=row, column=4).value = "Status"

    for col in range(1, 5):
        ws.cell(row=row, column=col).font = Font(bold=True)

    row += 1

    # ---------------------------------------
    # Analysis Data
    # ---------------------------------------

    for parameter, details in analysis.items():

        ws.cell(row=row, column=1).value = parameter
        ws.cell(row=row, column=2).value = details["value"]
        ws.cell(row=row, column=3).value = details["unit"]
        ws.cell(row=row, column=4).value = details["status"]

        row += 1

    wb.save(filename)

    return filename